#!/usr/bin/env python3
"""
Fase 1 del importador: lee el libro de Excel del negocio y lo convierte en JSON.

Se separa de la carga a propósito. Parsear es donde están las sorpresas —hojas con
formato irregular, filas de totales mezcladas con datos, nombres que traen datos
adentro— y conviene poder iterar sobre eso sin tocar la base de datos.

    python3 scripts/parse_workbook.py "example info/Pedidos2 copy.xlsm"

Escribe local/import-payload.json, que está gitignoreado porque lleva nombres,
direcciones y saldos de personas reales.

Sin deduplicación: es decisión explícita del negocio. Si un cliente aparece dos
veces escrito distinto, entra dos veces y se limpia a mano después.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl.  pip3 install openpyxl")

# Columnas de la hoja de clientes fijos que son productos.
# 'Pago' y 'Totales' viven entre ellas pero no lo son.
PRODUCT_COLUMNS = {
    "moras": "Moras",
    "mermelada": "Mermelada",
    "huevos": "Huevos",
    "miel": "Miel",
    "mandarinas": "Mandarinas",
    "limones": "Limones",
}

# Filas de resumen al pie de las hojas: no son clientes.
SUMMARY_ROWS = {"total", "$$", "libras totales de mora", "libras totales de moras"}


def norm(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def is_summary_row(name: str) -> bool:
    return norm(name) in SUMMARY_ROWS


def parse_contacts(ws) -> list[dict]:
    """Hoja CONTACTOS: nombre en la columna A, dirección en la B. Sin encabezado."""
    out = []
    for name, address, *_ in ws.iter_rows(values_only=True):
        if name is None or not str(name).strip():
            continue
        if is_summary_row(name):
            continue
        out.append(
            {
                "name": str(name).strip(),
                "address": str(address).strip() if address else None,
            }
        )
    return out


def parse_standing(ws) -> tuple[list[dict], dict]:
    """
    Hoja de clientes fijos: el pedido habitual de cada uno.

    Devuelve los clientes y el mapa de precios de la fila de precios del pie,
    que es de donde el negocio saca hoy cuánto vale cada cosa.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}

    header = [norm(c) for c in rows[0]]
    col_of = {}
    for key, product in PRODUCT_COLUMNS.items():
        if key in header:
            col_of[product] = header.index(key)

    customers = []
    price_row = None

    for row in rows[1:]:
        name = row[0] if row else None
        if name is None or not str(name).strip():
            # La fila de precios va suelta bajo la de totales, sin etiqueta.
            if price_row is None and any(
                isinstance(c, (int, float)) for c in row[2:8] if c is not None
            ):
                price_row = row
            continue

        if is_summary_row(name):
            continue

        items = []
        for product, idx in col_of.items():
            qty = row[idx] if idx < len(row) else None
            if isinstance(qty, (int, float)) and qty > 0:
                items.append({"product": product, "quantity": float(qty)})

        customers.append(
            {
                "name": str(name).strip(),
                "address": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                "standing_items": items,
            }
        )

    prices = {}
    if price_row:
        for product, idx in col_of.items():
            value = price_row[idx] if idx < len(price_row) else None
            if isinstance(value, (int, float)) and value > 0:
                prices[product] = int(value)

    return customers, prices


def find_month_sheet(wb) -> str | None:
    """Hoja de cartera mensual: empieza en 'Nombre' y termina en 'Debe'."""
    for name in wb.sheetnames:
        ws = wb[name]
        header = [norm(c) for c in next(ws.iter_rows(max_row=1, values_only=True), ())]
        if header and header[0] == "nombre" and "debe" in header:
            return name
    return None


def parse_balances(ws) -> list[dict]:
    """Columna 'Debe' de la hoja mensual: lo que cada cliente arrastra sin pagar."""
    rows = list(ws.iter_rows(values_only=True))
    header = [norm(c) for c in rows[0]]
    debe_idx = header.index("debe")

    out = []
    for row in rows[1:]:
        name = row[0] if row else None
        if name is None or not str(name).strip() or is_summary_row(name):
            continue
        amount = row[debe_idx] if debe_idx < len(row) else None
        if isinstance(amount, (int, float)) and amount > 0:
            out.append({"name": str(name).strip(), "balance_cop": int(round(amount))})
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="ruta al .xlsm del negocio")
    ap.add_argument(
        "--cutoff",
        required=True,
        help="fecha de corte de los saldos iniciales, YYYY-MM-DD. "
        "Es la fecha desde la que el FIFO cuenta la antigüedad de esa deuda.",
    )
    ap.add_argument("--out", default="local/import-payload.json")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)

    contacts_sheet = next((s for s in wb.sheetnames if norm(s) == "contactos"), None)
    standing_sheet = next((s for s in wb.sheetnames if norm(s) == "fijos"), None)
    month_sheet = find_month_sheet(wb)

    if not standing_sheet:
        return print("No encontré la hoja de clientes fijos.") or 1

    contacts = parse_contacts(wb[contacts_sheet]) if contacts_sheet else []
    standing, prices = parse_standing(wb[standing_sheet])
    balances = parse_balances(wb[month_sheet]) if month_sheet else []

    # Un cliente por nombre exacto. Los fijos mandan: traen pedido habitual.
    by_name: dict[str, dict] = {}

    for c in contacts:
        by_name[c["name"]] = {
            "name": c["name"],
            "address": c["address"],
            "recurrence": "ocasional",
            "standing_items": [],
            "opening_balance_cop": 0,
        }

    for c in standing:
        existing = by_name.get(c["name"])
        if existing:
            existing["address"] = existing["address"] or c["address"]
            existing["recurrence"] = "semanal"
            existing["standing_items"] = c["standing_items"]
        else:
            by_name[c["name"]] = {
                "name": c["name"],
                "address": c["address"],
                "recurrence": "semanal",
                "standing_items": c["standing_items"],
                "opening_balance_cop": 0,
            }

    unmatched = []
    for b in balances:
        target = by_name.get(b["name"])
        if target:
            target["opening_balance_cop"] = b["balance_cop"]
        else:
            unmatched.append(b["name"])

    customers = list(by_name.values())
    payload = {
        "cutoff_date": args.cutoff,
        "list_prices": prices,
        "customers": customers,
        "stats": {
            "contacts_sheet": contacts_sheet,
            "standing_sheet": standing_sheet,
            "month_sheet": month_sheet,
            "customers_total": len(customers),
            "customers_recurring": sum(
                1 for c in customers if c["recurrence"] == "semanal"
            ),
            "customers_with_balance": sum(
                1 for c in customers if c["opening_balance_cop"] > 0
            ),
            "balance_total_cop": sum(c["opening_balance_cop"] for c in customers),
            "standing_items_total": sum(len(c["standing_items"]) for c in customers),
            "balances_unmatched": len(unmatched),
        },
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), "utf-8")

    s = payload["stats"]
    print(f"Hojas leídas: {s['contacts_sheet']} · {s['standing_sheet']} · {s['month_sheet']}")
    print(f"Clientes:           {s['customers_total']}")
    print(f"  fijos semanales:  {s['customers_recurring']}")
    print(f"  con saldo:        {s['customers_with_balance']}")
    print(f"Líneas de pedido fijo: {s['standing_items_total']}")
    print(f"Saldo inicial total:   ${s['balance_total_cop']:,}".replace(",", "."))
    print(f"Precios de lista:      {prices}")

    if unmatched:
        # Cada uno de estos es plata que se queda sin cobrar si nadie lo mira.
        print(
            f"\n⚠️  {len(unmatched)} clientes con saldo NO existen en las otras hojas.\n"
            "    Se van a crear igual, pero revisa si son duplicados mal escritos."
        )

    print(f"\nEscrito en {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
